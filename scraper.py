import requests
from bs4 import BeautifulSoup
import pandas as pd

url = "https://books.toscrape.com/"

response = requests.get(url)

soup = BeautifulSoup(response.text, "html.parser")

print(soup.title.text)

products = soup.find_all("article", class_="product_pod")

data = []

for product in products:
    name = product.h3.a["title"]
    price = product.find("p", class_="price_color").text
    availability = product.find("p", class_="instock").text.strip()
    rating = product.find("p", class_="star-rating")["class"][1]

    data.append({
        "Product": name,
        "Price": price,
        "Rating": rating,
        "Availability": availability
    })

df = pd.DataFrame(data)

print(df)

# Clean Price
df["Price"] = df["Price"].str.replace("Â£", "", regex=False).astype(float)

# Convert Rating words to numbers
rating_map = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5
}

df["Rating"] = df["Rating"].map(rating_map)

print(df.head())
print(df.dtypes)

# Business KPIs

total_products = len(df)
average_price = df["Price"].mean()
highest_price = df["Price"].max()
lowest_price = df["Price"].min()
average_rating = df["Rating"].mean()

print("\n--- Business Summary ---")
print("Total Products:", total_products)
print("Average Price:", round(average_price, 2))
print("Highest Price:", highest_price)
print("Lowest Price:", lowest_price)
print("Average Rating:", round(average_rating, 2))

# Scrape all pages

all_data = []

for page in range(1, 51):
    url = f"https://books.toscrape.com/catalogue/page-{page}.html"

    response = requests.get(url)

    if response.status_code != 200:
        break

    soup = BeautifulSoup(response.text, "html.parser")

    products = soup.find_all("article", class_="product_pod")

    for product in products:
        name = product.h3.a["title"]
        price = product.find("p", class_="price_color").text
        availability = product.find("p", class_="instock").text.strip()
        rating = product.find("p", class_="star-rating")["class"][1]

        all_data.append({
            "Product": name,
            "Price": price,
            "Rating": rating,
            "Availability": availability
        })

print("Total products scraped:", len(all_data))

# Create final DataFrame from all scraped data

df = pd.DataFrame(all_data)

print("\nTotal rows:", len(df))
print(df.head())

# Create Clean Data

clean_df = df.copy()

# Clean Price
clean_df["Price"] = (
    clean_df["Price"]
    .str.replace("Â£", "", regex=False)
    .astype(float)
)

# Convert Rating into numbers
rating_map = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5
}

clean_df["Rating"] = clean_df["Rating"].map(rating_map)

# Clean Availability
clean_df["Availability"] = clean_df["Availability"].str.strip()

# Remove duplicate products
clean_df = clean_df.drop_duplicates(subset=["Product"])

print("\n--- Clean Data ---")
print(clean_df.head())

print("\nRows after cleaning:", len(clean_df))

print("\nData Types:")
print(clean_df.dtypes)

# KPI Analysis

total_products = len(clean_df)
average_price = clean_df["Price"].mean()
highest_price = clean_df["Price"].max()
lowest_price = clean_df["Price"].min()
average_rating = clean_df["Rating"].mean()

in_stock = clean_df["Availability"].str.contains("In stock").sum()
out_of_stock = total_products - in_stock

print("\n--- KPI Summary ---")
print("Total Products:", total_products)
print("Average Price:", round(average_price, 2))
print("Highest Price:", round(highest_price, 2))
print("Lowest Price:", round(lowest_price, 2))
print("Average Rating:", round(average_rating, 2))
print("In Stock:", in_stock)
print("Out of Stock:", out_of_stock)

# Analysis

# Top 10 highest-priced products
top_expensive = clean_df.nlargest(10, "Price")[
    ["Product", "Price", "Rating", "Availability"]
]

# Top 10 highest-rated products
top_rated = clean_df.nlargest(10, "Rating")[
    ["Product", "Price", "Rating", "Availability"]
]

# Number of products by rating
rating_analysis = (
    clean_df.groupby("Rating")
    .agg(
        Product_Count=("Product", "count"),
        Average_Price=("Price", "mean")
    )
    .reset_index()
)

# Products by availability
stock_analysis = (
    clean_df.groupby("Availability")
    .agg(
        Product_Count=("Product", "count"),
        Average_Price=("Price", "mean")
    )
    .reset_index()
)

print("\n--- Top 10 Expensive Products ---")
print(top_expensive)

print("\n--- Top 10 Rated Products ---")
print(top_rated)

print("\n--- Rating Analysis ---")
print(rating_analysis)

print("\n--- Stock Analysis ---")
print(stock_analysis)

# Export Final Report to Excel

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

output_file = "Marketplace_Performance_Report.xlsx"

# KPI DataFrame
kpi_df = pd.DataFrame({
    "KPI": [
        "Total Products",
        "Average Price",
        "Highest Price",
        "Lowest Price",
        "Average Rating",
        "In Stock",
        "Out of Stock"
    ],
    "Value": [
        total_products,
        round(average_price, 2),
        round(highest_price, 2),
        round(lowest_price, 2),
        round(average_rating, 2),
        in_stock,
        out_of_stock
    ]
})

# Create Excel file
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    df.to_excel(writer, sheet_name="Raw Data", index=False)

    clean_df.to_excel(writer, sheet_name="Clean Data", index=False)

    kpi_df.to_excel(writer, sheet_name="KPI", index=False)

    # Put different analysis tables into one sheet
    top_expensive.to_excel(
        writer,
        sheet_name="Analysis",
        index=False,
        startrow=1
    )

    top_rated.to_excel(
        writer,
        sheet_name="Analysis",
        index=False,
        startrow=15
    )

    rating_analysis.to_excel(
        writer,
        sheet_name="Analysis",
        index=False,
        startrow=29
    )

    stock_analysis.to_excel(
        writer,
        sheet_name="Analysis",
        index=False,
        startrow=38
    )

# Format workbook
wb = load_workbook(output_file)

for ws in wb.worksheets:

    # Format header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

# Add section titles to Analysis sheet
analysis_ws = wb["Analysis"]

analysis_ws["A1"] = "Top 10 Highest-Priced Products"
analysis_ws["A15"] = "Top 10 Highest-Rated Products"
analysis_ws["A29"] = "Rating Analysis"
analysis_ws["A38"] = "Stock Analysis"

for cell in ["A1", "A15", "A29", "A38"]:
    analysis_ws[cell].font = Font(bold=True)

wb.save(output_file)

print("\nExcel report created successfully!")
print("File:", output_file)